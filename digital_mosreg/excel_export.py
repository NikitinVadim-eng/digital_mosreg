from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from digital_mosreg.flatten import COLUMN_ORDER, rows_to_column_order

LINK_COLUMNS = frozenset({"URL", "Ссылка на проект"})


def _write_cell(cell: Cell, column: str, value: Any) -> None:
    text = "" if value is None else str(value)
    if column in LINK_COLUMNS and text.startswith(("http://", "https://")):
        cell.value = text
        cell.hyperlink = text
        cell.font = Font(color="0563C1", underline="single")
    else:
        cell.value = value if not isinstance(value, str) else text


def rows_to_workbook(rows: list[dict[str, Any]]) -> Workbook:
    ordered = rows_to_column_order(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ИИ-решения"
    for col_idx, name in enumerate(COLUMN_ORDER, start=1):
        sheet.cell(row=1, column=col_idx, value=name)
    for row_idx, row in enumerate(ordered, start=2):
        for col_idx, name in enumerate(COLUMN_ORDER, start=1):
            _write_cell(sheet.cell(row=row_idx, column=col_idx), name, row.get(name, ""))
    for col_idx in range(1, len(COLUMN_ORDER) + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18
    return workbook


def rows_to_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    workbook = rows_to_workbook(rows)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_rows_xlsx(rows: list[dict[str, Any]], path: Path | str) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = rows_to_workbook(rows)
    workbook.save(output)
    return output
